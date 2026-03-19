import openpyxl
import datetime
import re

# ファイルパスの設定
input_file = "/Users/harumacmini/programming/info_companyDetail/data/template/【マテリアル】_デビス㈱.xlsx"
output_file = "/Users/harumacmini/programming/info_companyDetail/data/template/【マテリアル】_テンプレート.xlsx"

# =========================================================
# 【可変設定エリア】
# =========================================================

# ① 年月（日付）セルを置き換えるテンプレート文字列
# ⇒ 決算期ヘッダーなどを削除せず、この文字列に一括変換します。自由に書き換えてください。
DATE_TEMPLATE = "20XX年XX月期"

# ② 文字列や独自の科目名などを強制クリアする範囲
CLEAR_TARGETS = {
    "基本情報": ["C10:C30", "G10:G30"],          
    "従業員・資格一覧": ["F8:I100"],             
    "役員一覧": ["C4:D20", "E4:E20", "I4:I20"],  
    "株主一覧": ["B7:D30", "G7:G30"],            
    "土地": ["C7:C30", "K7:K30"],                
    "車輛一覧": ["B7:D30", "H7:H30", "J7:J30"],  
    "退職金": ["E7:E50"],
}
# =========================================================


def is_date_like(value):
    """セルが日付・年月っぽいかどうかを判定します"""
    if value is None:
        return False
    # Excelのシリアル値（日付データとして認識されているもの）
    if isinstance(value, datetime.datetime):
        return True
    # 文字列でベタ打ちされている場合（例: "2024-12-31", "2024年12月期", "2024/12" など）
    if isinstance(value, str):
        val_str = value.strip()
        # 1900年代か2000年代で始まり、ハイフン/スラッシュ/年/月などが含まれるパターン
        pattern = r'^(19|20)\d{2}[-/年]\d{1,2}([-/月]\d{1,2}日?)?(期)?$'
        if re.match(pattern, val_str):
            return True
    return False

def is_number_like(value):
    """セルの中身が「数値」または「文字列化された数値」かを判定します"""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        val_str = value.replace(',', '').replace(' ', '').replace('¥', '').replace('￥', '').strip()
        if val_str == '-':
            return True
        try:
            float(val_str)
            return True
        except ValueError:
            pass
    return False

print("ファイルを読み込んでいます... (少し時間がかかります)")
wb = openpyxl.load_workbook(input_file, data_only=False)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # ---------------------------------------------------------
    # ① シート全体の「年月」「数値」を自動判定して処理
    # ---------------------------------------------------------
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
                
            # 数式(Formula)は絶対に保護する
            if cell.data_type == 'f' or str(cell.value).startswith('='):
                continue
            
            # 【新規】年月・日付データなら、テンプレート文字列に置換
            if is_date_like(cell.value):
                cell.value = DATE_TEMPLATE
                continue
            
            # 数値、または文字列化された数値ならクリア
            if is_number_like(cell.value):
                cell.value = None

    # ---------------------------------------------------------
    # ② CLEAR_TARGETS で指定された範囲の「文字列」を強制クリア
    # ---------------------------------------------------------
    if sheet_name in CLEAR_TARGETS:
        for cell_range in CLEAR_TARGETS[sheet_name]:
            try:
                for row in ws[cell_range]:
                    for cell in row:
                        if cell.value is None:
                            continue
                        
                        # 数式でなければすべてクリア
                        if cell.data_type != 'f' and not str(cell.value).startswith('='):
                            cell.value = None
            except Exception as e:
                print(f"範囲スキップ ({sheet_name} - {cell_range}): {e}")

# 保存
print("テンプレートファイルを保存しています...")
wb.save(output_file)
print(f"完了しました！\n出力先: {output_file}")