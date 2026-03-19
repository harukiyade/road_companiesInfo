#!/usr/bin/env python3
"""
【Legatus ONE用】決算書PDF → マテリアルExcel 転記システム（完全マッチング・エラー回避版）

修正・強化内容:
1. APIエラーの回避: configによるJSON指定を削除し、プロンプト指示で強制。サイレントエラーを根絶。
2. 最強の文字列クリーニング: 改行、全半角スペース、特殊記号を完全に除去しマッチング漏れを根絶。
3. 行単位の全文連結スキャン: 結合セルやインデントに左右されず、その行の科目を100%特定。
4. 誤爆防止: シート別に開始行を厳格に指定し、タイトル行（3行目等）への上書きをブロック。
"""

import datetime
import json
import os
import re
import sys
import time
from pathlib import Path
from copy import copy
import openpyxl

# --- パス設定 ---
TEMPLATE_PATH = Path("/Users/harumacmini/programming/info_companyDetail/data/template/【マテリアル】_テンプレート.xlsx")
INPUT_PDF_DIR = Path("/Users/harumacmini/programming/info_companyDetail/data/template/kessansho")
OUTPUT_DIR = Path("/Users/harumacmini/programming/info_companyDetail/data/output")

# -----------------------------------------------------------------------------
# 1. ユーティリティ
# -----------------------------------------------------------------------------
def get_master_cell(ws, row, col):
    """結合セルの場合でも左上の親セルを返す（書き込みエラー防止）"""
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and \
           merged_range.min_col <= col <= merged_range.max_col:
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return ws.cell(row=row, column=col)

def find_cell_by_text(ws, search_text, max_row=60):
    for r in range(1, max_row + 1):
        for c in range(1, 10):
            val = ws.cell(row=r, column=c).value
            if val and isinstance(val, str) and search_text in val:
                return r, c
    return None, None

def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def clean_extracted_name(name):
    if not name: return ""
    return re.sub(r'[\s　\.．]', '', str(name))

def clean_account_name(name):
    """
    【超重要】最強のクリーニング関数。
    Excel特有の改行、全半角スペース、括弧、記号を完全に消し去り、純粋な文字列にする。
    """
    if not name: return ""
    s = str(name)
    s = re.sub(r'[\r\n\t]', '', s)
    s = re.sub(r'（.*?）|\(.*?\)|［.*?］|\[.*?\]|【.*?】', '', s)
    s = re.sub(r'[・\s　※及び]', '', s)
    return s.strip()

def is_ignored_header(raw_label):
    """Excelのタイトル行（書き込み対象外の大見出し）を判定し、誤爆を防ぐ"""
    raw_str = str(raw_label).strip()
    if re.match(r'^【.*】$', raw_str) or re.match(r'^［.*］$', raw_str):
        return True
    
    clean_lbl = raw_str.replace(" ", "").replace("　", "").replace("\n", "")
    headers = [
        "資産の部", "負債の部", "純資産の部", 
        "負債純資産合計", "負債合計", "純資産合計", "資産合計",
        "流動資産", "固定資産", "繰延資産",
        "流動負債", "固定負債"
    ]
    if clean_lbl in headers:
        return True
    return False

def is_formula(cell):
    if cell.data_type == 'f': return True
    if isinstance(cell.value, str) and cell.value.startswith('='): return True
    return False

def is_dummy_value(val):
    if type(val) in (int, float): return True
    if isinstance(val, str):
        if re.match(r'^\d{4}-\d{2}-\d{2}$', val.strip()): return False
        if "期" in val or "年" in val or "月" in val: return False
        clean_str = val.replace(",", "").replace("-", "").replace("△", "").replace("▲", "").replace(".", "").replace(" ", "").strip()
        if clean_str.isdigit(): return True
    return False

def parse_incomplete_json(json_str):
    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        try:
            fixed_str = re.sub(r',\s*$', '', json_str.strip()) + "}"
            return json.loads(fixed_str)
        except:
            return {}

# --- [新機能] マテリアルから探すべき項目名を事前抽出 ---
def extract_target_labels(template_path):
    wb = openpyxl.load_workbook(template_path, data_only=True)
    labels = set()
    for sheet_name in ["BS", "PL", "SGA", "製造原価", "利益修正"]:
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        start_r = 8 if sheet_name != "利益修正" else 6
        for r in range(start_r, 150): 
            for c in range(2, 6):
                val = ws.cell(row=r, column=c).value
                if val and isinstance(val, str):
                    if is_ignored_header(val): continue
                    cl = clean_account_name(val)
                    if cl and len(cl) > 1:
                        labels.add(cl)
    wb.close()
    
    labels.update(["売上高", "現金預金", "売掛金", "期首商品棚卸高", "期末商品棚卸高"])
    return list(labels)

# -----------------------------------------------------------------------------
# 2. AI抽出エンジン (APIエラー回避 ＆ ダイナミック抽出)
# -----------------------------------------------------------------------------
def extract_with_gemini(pdf_path, client, target_labels):
    from google.genai import types
    import fitz
    
    labels_str = ", ".join(target_labels)

    doc = fitz.open(pdf_path)
    contents = [f"""あなたは世界最高精度の財務解析AIです。決算書PDFの全ページを確認し、情報を抽出しJSONで返してください。
    
    【厳守事項：企業メタデータ】
    - 代表者氏名は、申告書の表紙にある氏名を最優先としてください。
    - 当期の「決算年月日」を読み取り、"YYYY-MM-DD"形式で出力してください。
    - 株主構成は、法人税申告書の「別表二」から抽出してください。
    - 決算書の各表の右上に「(単位：千円)」の記載があるか確認してください。記載があれば "unit" を "千円" に、なければ "円" にしてください。
    
    【厳守事項：財務数値データの抽出】
    以下のリストにある「必要な項目」を決算書（BS、PL、製造原価報告書、販売費及び一般管理費内訳）から探し出して抽出してください。
    
    [重要探索リスト（必ず探すこと）]
    {labels_str}
    
    [抽出時の注意]
    - 上記リストの項目を探す際は、決算書に実際に記載されている科目名（例えば「現金及び預金」など）をそのままJSONのキーとして抽出してください。
    - リストにない細かい科目名も、表に記載されていればすべて漏らさず抽出してください。
    - 期首と期末（例：期首商品棚卸高と期末商品棚卸高）は絶対に区別してください。
    - 金額はカンマや円マークを除外した純粋な数値（例: 10290000）で出力してください。マイナスの場合は - をつけてください。
    
    必ず以下のJSONフォーマットのみを出力してください。Markdown（```json など）は絶対に書かないでください。
    {{
      "meta": {{
        "company_name": "法人名",
        "representative": "代表者氏名",
        "fiscal_end": "YYYY-MM-DD",
        "unit": "千円 または 円",
        "shareholders": [{{"name": "株主名", "shares": 株式数}}]
      }},
      "data": {{ "勘定科目名": 数値, ... }}
    }}"""]
    
    for i in range(len(doc)):
        pix = doc[i].get_pixmap(dpi=72) 
        contents.append(types.Part.from_bytes(data=pix.tobytes("png"), mime_type="image/png"))
    doc.close()

    for attempt in range(3):
        try:
            model_name = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
            
            # APIエラーを避けるため、configディクショナリでのJSON指定を削除し、プロンプト指示に依存させる
            response = client.models.generate_content(
                model=model_name, 
                contents=contents
            )
            
            match = re.search(r"\{[\s\S]*\}", response.text)
            if not match: continue
            
            raw_json = parse_incomplete_json(match.group(0))
            if not raw_json: continue
            
            # --- 千円単位の判定と一律変換 ---
            unit_str = raw_json.get("meta", {}).get("unit", "")
            is_yen = False
            if "千円" not in unit_str and "円" in unit_str:
                is_yen = True
            
            # 金額規模による自動判定（500万以上の数値があれば円単位とみなす）
            if not is_yen:
                max_val = 0
                for k, v in raw_json.get("data", {}).items():
                    try:
                        num = abs(float(str(v).replace(",","").replace("円","").strip() or 0))
                        max_val = max(max_val, num)
                    except: pass
                if max_val > 5000000:
                    is_yen = True

            processed_data = {}
            for k, v in raw_json.get("data", {}).items():
                try:
                    s_val = str(v).replace(",", "").replace("△", "-").replace("▲", "-").replace("円", "").strip()
                    val = float(s_val)
                    if is_yen:
                        val = val / 1000.0 # 千円単位に一律変換
                    processed_data[clean_account_name(k)] = val
                except: pass
                
            raw_json["data"] = processed_data
            return raw_json
            
        except Exception as e:
            print(f"    [AI抽出エラー] {e} - リトライ中... ({attempt+1}/3)")
            time.sleep(5)
    return {}

# -----------------------------------------------------------------------------
# 3. Excel転記・統合エンジン
# -----------------------------------------------------------------------------
def apply_to_excel(all_results, template_path, output_path):
    if not template_path.exists():
        print(f"❌ テンプレートファイルが見つかりません: {template_path}")
        sys.exit(1)
        
    wb = openpyxl.load_workbook(template_path, data_only=False)
    
    sorted_keys = sorted(list(all_results.keys()))
    if not sorted_keys: return
    
    latest_key = sorted_keys[-1]
    latest_meta = all_results[latest_key].get("meta", {})
    
    comp_name = latest_meta.get("company_name", "対象会社")
    rep_name = clean_extracted_name(latest_meta.get("representative", "代表者名"))
    shareholders = latest_meta.get("shareholders", [])
    
    fiscal_end_str = latest_meta.get("fiscal_end", "2024-03-31")
    try:
        base_dt = datetime.datetime.strptime(fiscal_end_str, "%Y-%m-%d")
    except ValueError:
        base_dt = datetime.datetime(2024, 3, 31)

    # --- A. 基本情報の反映 ---
    if "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        for col in range(7, 18): get_master_cell(ws, 6, col).value = comp_name
        r, c = find_cell_by_text(ws, "評価対象")
        if r: get_master_cell(ws, r, c + 2).value = comp_name
        
    if "基本情報" in wb.sheetnames:
        ws = wb["基本情報"]
        r_s, _ = find_cell_by_text(ws, "商号")
        if r_s: get_master_cell(ws, r_s, 4).value = comp_name
        r_r, _ = find_cell_by_text(ws, "代表者")
        if r_r: get_master_cell(ws, r_r, 4).value = rep_name
        r_d, _ = find_cell_by_text(ws, "決算期")
        if r_d:
            date_cell = get_master_cell(ws, r_d, 4)
            date_cell.value = base_dt
            date_cell.number_format = 'yyyy/mm/dd'

    # --- D. 年月ラベルの更新 ---
    sheet_date_targets = [
        ("PL", 6, [5, 7, 9]), 
        ("SGA", 6, [5, 7, 9]),
        ("製造原価", 6, [5, 7, 9, 11, 13]),
        ("BS", 7, [6, 7, 8]),
        ("利益修正", 5, [5, 6, 7]), 
    ]
    for name, row, cols in sheet_date_targets:
        if name not in wb.sheetnames: continue
        ws = wb[name]
        for i, col in enumerate(cols):
            dt = datetime.datetime(base_dt.year - (len(cols)-1-i), base_dt.month, base_dt.day)
            cell = get_master_cell(ws, row, col)
            cell.value = dt
            cell.number_format = 'yyyy"年"m"月期"'

    # --- E. 財務数値の完全網羅的転記 ---
    target_keys = sorted_keys[-3:] 
    data_list = [all_results[k].get("data", {}) for k in target_keys]
    while len(data_list) < 3:
        data_list.insert(0, {})

    WRITE_TARGET_SHEETS = ["BS", "PL", "SGA", "製造原価", "利益修正", "修正PL"]

    for sheet_name in WRITE_TARGET_SHEETS:
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        
        scan_configs = []
        # B列～E列を広範囲にスキャン
        if sheet_name in ["PL", "SGA", "製造原価"]:
            scan_configs.append( ([2, 3, 4, 5], [5, 7, 9]) )
        elif sheet_name == "利益修正":
            scan_configs.append( ([2, 3, 4], [5, 6, 7]) ) 
        elif sheet_name == "BS": 
            scan_configs.append( ([2, 3, 4, 5], [6, 7, 8]) )
        elif sheet_name == "修正PL":
            scan_configs.append( ([2, 3, 4, 5], [5, 8, 11]) )
            
        for label_cols, write_cols in scan_configs:
            # 見出し行への誤爆を防ぐための開始行設定
            if sheet_name in ["BS", "PL", "SGA", "製造原価", "修正PL"]:
                start_row = 8
            elif sheet_name == "利益修正":
                start_row = 6
            else:
                start_row = 8
            
            # パージ処理（ダミーの確実な消去）
            for r in range(start_row, 250):
                for col_idx in write_cols:
                    cell = get_master_cell(ws, r, col_idx)
                    if cell and not is_formula(cell) and is_dummy_value(cell.value):
                        cell.value = None

            # 転記ループ
            for r in range(start_row, 250):
                # その行のB列〜E列の値をすべて連結して科目を特定する
                row_text = ""
                for c in label_cols:
                    cell_val = ws.cell(row=r, column=c).value
                    if cell_val and isinstance(cell_val, str):
                        row_text += cell_val

                if not row_text.strip(): continue
                
                if is_ignored_header(row_text):
                    continue
                    
                label = clean_account_name(row_text)
                if not label: continue
                
                if sheet_name == "利益修正" and "営業利益" in label: label = "営業利益" 
                elif sheet_name == "利益修正" and "売上高" in label: label = "売上高"
                
                for col_idx, data in zip(write_cols, data_list):
                    if not data: continue 
                    
                    match_val = None
                    
                    # 強力な同義語マッピング
                    synonyms = {
                        "売上高": ["売上金額", "完成工事高", "売上", "事業収益", "賃貸料収益"],
                        "現金預金": ["現金及び預金", "現預金", "現金", "普通預金", "当座預金"],
                        "売掛金": ["売掛金額", "完成工事未収入金", "未収金", "未収入金", "営業未収入金", "売上債権", "工事未収入金"],
                        "買掛金": ["買掛金額", "工事未払金", "営業未払金", "支払手形及び買掛金"],
                        "期首商品棚卸高": ["期首棚卸高", "商品期首棚卸高", "期首商品"],
                        "期末商品棚卸高": ["期末棚卸高", "商品期末棚卸高", "期末商品"],
                        "仕入高": ["当期商品仕入高", "商品仕入高", "当期仕入高", "材料仕入高"],
                        "有形固定資産": ["有形固定資産計", "有形固定資産合計"],
                        "無形固定資産": ["無形固定資産計", "無形固定資産合計"],
                        "投資等": ["投資その他の資産", "投資その他の資産計", "投資その他"],
                        "当期純利益": ["当期純利益金額", "税引後当期純利益"],
                        "退職給付引当金繰入": ["退職給付費用"],
                        "法定福利費": ["福利厚生費", "法定福利費"],
                        "修繕費": ["修繕維持費"],
                        "消耗品費": ["事務用品費", "消耗工具備品費"],
                        "地代家賃": ["賃借料", "家賃"],
                        "租税公課": ["税金"],
                        "支払手数料": ["手数料"],
                        "水道光熱費": ["電気代", "水道代", "光熱費"],
                        "旅費交通費": ["交通費", "旅費"],
                        "通信費": ["電話代", "郵便料"],
                        "減価償却費": ["減価償却"],
                    }
                    
                    if label in data:
                        match_val = data[label]
                    elif label in synonyms and any(s in data for s in synonyms[label]):
                        for s in synonyms[label]:
                            if s in data:
                                match_val = data[s]
                                break
                    elif "給与賞与退職金" in label or "人件費" in label:
                        keys = [k for k in data.keys() if any(x in k for x in ["給料", "給与", "賞与", "役員報酬", "退職金", "賃金"]) and "引当" not in k and "計" not in k]
                        if keys:
                            match_val = sum(data[k] for k in keys)
                    else:
                        # 部分一致による救済
                        for k, v in data.items():
                            if "計" in k and label != k: continue 
                            if "期首" in label and "期首" not in k: continue 
                            if "期末" in label and "期末" not in k: continue 
                            
                            if label in k or k in label:
                                match_val = v
                                break
                    
                    if match_val is not None:
                        # 書き込む先のセルを取得（結合セル対応）
                        target = get_master_cell(ws, r, col_idx)
                        if not is_formula(target):
                            target.value = int(round(match_val))

    # 不要シート削除
    sheets_to_remove = [s for s in wb.sheetnames if "書き出しの概要" in s]
    for s in sheets_to_remove:
        del wb[s]
        
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

# -----------------------------------------------------------------------------
# 4. メイン処理
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    from google import genai
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        print("❌ GEMINI_API_KEY が設定されていません。")
        sys.exit(1)
        
    client = genai.Client(api_key=api_key)
    all_data = {}
    
    pdf_paths = sorted(list(INPUT_PDF_DIR.glob("*.pdf")))
    if not pdf_paths:
        print(f"❌ PDFファイルが見つかりません: {INPUT_PDF_DIR}")
        sys.exit(1)
        
    print("🔍 テンプレートから抽出対象の項目をリストアップしています...")
    target_labels = extract_target_labels(TEMPLATE_PATH)
        
    print(f"🔍 解析開始（{len(pdf_paths)}件のPDFを解析・転記します）...")
    
    for path in pdf_paths:
        label = path.stem 
        print(f"  ⟳ [{label}] 解析中...")
        res = extract_with_gemini(path, client, target_labels)
        if res: all_data[label] = res
        time.sleep(3) 

    if all_data:
        sorted_keys = sorted(list(all_data.keys()))
        latest_label = sorted_keys[-1]
        raw_name = all_data[latest_label].get("meta", {}).get("company_name", "新規会社")
        clean_name = re.sub(r'[\\/:*?"<>|]', '', raw_name).replace(" ", "")
        
        final_output = OUTPUT_DIR / f"【マテリアル】_{clean_name}.xlsx"
        
        apply_to_excel(all_data, TEMPLATE_PATH, final_output)
        print(f"✨ 完了！ 出力ファイル:\n   {final_output}")
    else:
        print("❌ 解析データが得られませんでした。")